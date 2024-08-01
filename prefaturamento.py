import io
import streamlit as st
import pandas as pd
import datetime
#import numpy as np
#import matplotlib.pyplot as plt
from io import StringIO

# Import statistics Library
#import statistics
from openpyxl import reader,load_workbook,Workbook
#import xlsxwriter
#from openpyxl.utils.dataframe import dataframe_to_rows

PAGE_TITLE = "App de Pré-Faturamento"
PAGE_ICON = "https://cse.energisa.com.br/img/grupo_energisa.png"
st.set_page_config(
    page_title=PAGE_TITLE,
    page_icon=PAGE_ICON,
    layout="wide",
    menu_items={
        'About': "Bugs ou sugestões, enviar um e-mail para joebert.maia@energisa.com.br"
    }
)
st.markdown("<img src='https://www.grupoenergisa.com.br/sites/default/files/Logo.svg' width='200' style='display: block; margin: 0 auto;'>" , unsafe_allow_html=True)
st.markdown("<h1 style='text-align: center;'>Pré-faturamento</h1>", unsafe_allow_html=True)
st.markdown("<h5 style='text-align: center;'>Webapp para o processamento das unidades do grupo A ou grupo B que precisam ser enviadas para manutenção.</h1>", unsafe_allow_html=True)
st.markdown("""<style>[aria-label="dialog"]{width: 90%;}</style>""", unsafe_allow_html=True)

hide_streamlit_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            #stToolbar {visibility: hidden;}
            </style>
            """
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

#st.title('Pré-faturamento')
#st.write("Esse webapp auxilia no envio das listas de unidades para manutenção do grupo A e grupo B.")

uploaded_file = st.file_uploader("Escolha uma planilha", type = 'xlsx')

wb = Workbook()

if uploaded_file is not None:
    wb = load_workbook(uploaded_file, read_only=False)
    st.write(f"Nome do arquivo: {uploaded_file.name}")
    st.write(f"Nome das planilhas: {wb.sheetnames}")

    prefat = pd.read_excel(uploaded_file, engine='openpyxl',skiprows=2)
    st.write(prefat.head())

    filtro_desconectada = prefat[prefat['SITUAÇÃO_COMUNICAÇÃO'] == 'DESCONECTADA']
    st.write('UCs desconectadas')
    st.write(filtro_desconectada)

buffer = io.BytesIO()
wb.save(buffer)

st.download_button(
label="Download da planilha do Excel",
data=buffer,
file_name="fk.xlsx",
)





footer="""<style>

.footer {
position: absolute;
top:230px;
left: 0;
bottom: 0px;
width: 100%;
background-color: transparent;
color: filter: invert(1); black;
text-align: center;
}
</style>
<div class="footer">
<hr style='width:70%;text-align:center;margin:auto'>
<p>Centro de Operação da Medição (COM) - Grupo Energisa <br> <a href="mailto:com@energisa.com.br">com@energisa.com.br</a></p>
</div>
"""
st.markdown(footer,unsafe_allow_html=True)
